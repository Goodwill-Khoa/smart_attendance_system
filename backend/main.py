from fastapi import FastAPI, Depends, HTTPException, Request, Header, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from sqlalchemy.orm import Session
from sqlalchemy import inspect, text
import models
import schemas
from authorization import auth, get_current_user
from database import engine, get_db
from models import UserRole
import uuid
import time
from datetime import datetime, timezone, timedelta
import hashlib
import io, csv
from fastapi.responses import Response
import os
import re
import requests
import secrets
import string
from pathlib import Path
from openpyxl import load_workbook

models.Base.metadata.create_all(bind=engine)
optional_security = HTTPBearer(auto_error=False)
DEFAULT_LECTURER_PASSWORD = os.getenv("DEFAULT_LECTURER_PASSWORD", "qwe123")
DEFAULT_TEST_STUDENT_PASSWORD = os.getenv("DEFAULT_TEST_STUDENT_PASSWORD", "qwe123")
SUPERUSER_USERNAME = os.getenv("SUPERUSER_USERNAME", "superuser")
SUPERUSER_PASSWORD = os.getenv("SUPERUSER_PASSWORD", "admin123")
SUPERUSER_EMAIL = os.getenv("SUPERUSER_EMAIL", "superuser@inf.elte.hu")
SUPERUSER_DISPLAY_NAME = os.getenv("SUPERUSER_DISPLAY_NAME", "Superuser")
PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEACHER_ROSTER_TEMPLATE_PATH = PROJECT_ROOT / "frontend" / "public" / "teacher_roster_template.csv"

def current_semester_label() -> str:
    now = datetime.now(timezone.utc)
    season = "Spring" if now.month <= 6 else "Autumn"
    return f"{season} {now.year}"

def ensure_schema_upgrades():
    inspector = inspect(engine)
    course_columns = {column["name"] for column in inspector.get_columns("courses")}
    user_columns = {column["name"] for column in inspector.get_columns("users")}
    if "semester" not in course_columns:
        with engine.begin() as connection:
            connection.execute(text("ALTER TABLE courses ADD COLUMN semester VARCHAR"))
            connection.execute(text("UPDATE courses SET semester = :semester WHERE semester IS NULL"), {"semester": current_semester_label()})
    if "course_type" not in course_columns:
        with engine.begin() as connection:
            connection.execute(text("ALTER TABLE courses ADD COLUMN course_type VARCHAR"))
            connection.execute(text("UPDATE courses SET course_type = 'L' WHERE course_type IS NULL"))
    if "is_admin" not in user_columns:
        with engine.begin() as connection:
            connection.execute(text("ALTER TABLE users ADD COLUMN is_admin BOOLEAN DEFAULT 0"))
            connection.execute(text("UPDATE users SET is_admin = 0 WHERE is_admin IS NULL"))

DEFAULT_COURSES = [
    ("Advanced Software Technology", "L", ""),
    ("Design and analysis of algorithms", "L", "IPM-22fDAAE"),
    ("Design and analysis of algorithms", "Pr", "IPM-22fDAAG"),
    ("Formal semantics", "L", "IPM-22feszFSE"),
    ("Formal semantics", "Pr", "IPM-22feszFSG"),
    ("Software quality and testing", "L", "IPM-22feszSQTE"),
    ("Software quality and testing", "Pr", "IPM-22feszSQTE"),
    ("Software Technology", "Lab I", "IPM-22feszSALAB1"),
    ("Machine Learning", "L", "IPM-22feszMLEE"),
    ("Machine Learning", "Pr", "IPM-22feszMLEG"),
    ("Service Science", "L", "IPM-22feszSESCE"),
    ("Service Science", "Pr", "IPM-22feszSESCE"),
]

DEFAULT_LECTURERS = [
    ("Gergely Dévai", "deva@inf.elte.hu", "Course Management, Lectures, Material"),
    ("István Bátori", "istvan.batori@gmail.com", "Coaching"),
    ("Khawla Bouafia", "bouafia@inf.elte.hu", "Coaching"),
    ("Dániel Leskó", "ldani@elte.hu", "Coaching"),
    ("Attila Ulbert", "attila.ulbert@gmail.com", "Coaching"),
    ("Bálint Orosz", "balint.orosz@inf.elte.hu", "Coaching (Substitute for Attila Ulbert)"),
    ("Pál György Sarmasági", "psarmasagi@inf.elte.hu", "Coaching"),
    ("Zsolt Pelyhe", "zsoltipelyhe@gmail.com", "Lecturer"),
    ("Attila Bence Buzgan", "ba@inf.elte.hu", "Sample Lecturer"),
]

FALLBACK_TEACHER_ROSTER_ACCOUNTS = [
    ("gvq37g@inf.elte.hu", "Bouafia Khawla"),
    ("i8yvmx@inf.elte.hu", "Buzgán Attila Bence"),
    ("pc1w7y@inf.elte.hu", "Bátori István"),
    ("eun3j8@inf.elte.hu", "Dévai Gergely"),
    ("q1p3pw@inf.elte.hu", "Leskó Dániel"),
    ("bck909@inf.elte.hu", "Orosz Bálint Dominik"),
    ("a6f41u@inf.elte.hu", "Pelyhe Zsolt"),
    ("rep4jh@inf.elte.hu", "Pásztor-Nagy Anett"),
    ("fykiq8@inf.elte.hu", "Sarmasági Pál György"),
    ("sa3owu@inf.elte.hu", "Ulbert Attila Dr."),
]

MENTAL_ACCESS_STUDENTS = [
    ("student1@inf.elte.hu", "John Doe"),
    ("student2@inf.elte.hu", "Jane Smith"),
    ("student3@inf.elte.hu", "Alice Johnson"),
    ("student4@inf.elte.hu", "Bob Wilson"),
    ("student5@inf.elte.hu", "Carol Brown"),
    ("student6@inf.elte.hu", "David Miller"),
    ("student7@inf.elte.hu", "Emma Davis"),
    ("student8@inf.elte.hu", "Frank Taylor"),
    ("student9@inf.elte.hu", "Grace Anderson"),
    ("student10@inf.elte.hu", "Henry Thomas"),
    ("student11@inf.elte.hu", "Ivy Jackson"),
    ("student12@inf.elte.hu", "Jack Martin"),
    ("a9n7p5@inf.elte.hu", "Goodwill Khoa"),
    ("lyjzux@inf.elte.hu", "Maxim Curos"),
    ("lq38sf@inf.elte.hu", "Xuetao Li"),
    ("ctzr62@inf.elte.hu", "Yacine Naat"),
    ("fvmi7v@inf.elte.hu", "Kevin Zsombók"),
]
MENTAL_ACCESS_STUDENT_MAP = {
    email.lower(): name for email, name in MENTAL_ACCESS_STUDENTS
}


def get_mental_access_student_name(email: str) -> str | None:
    return MENTAL_ACCESS_STUDENT_MAP.get((email or "").strip().lower())


def is_mental_access_student(email: str) -> bool:
    return get_mental_access_student_name(email) is not None


def get_uploaded_course_student_roster_map(db: Session, course_id: uuid.UUID) -> dict[str, str]:
    return {
        item.student_email.lower(): (item.student_name or "").strip()
        for item in db.query(models.CourseStudentRegistry)
        .filter(models.CourseStudentRegistry.course_id == course_id)
        .all()
    }


def get_effective_course_student_roster_map(db: Session, course_id: uuid.UUID) -> dict[str, str]:
    roster_map = get_uploaded_course_student_roster_map(db, course_id)
    for email, name in MENTAL_ACCESS_STUDENT_MAP.items():
        if not roster_map.get(email):
            roster_map[email] = name
    return roster_map


def load_teacher_roster_template_accounts() -> list[tuple[str, str]]:
    if not TEACHER_ROSTER_TEMPLATE_PATH.exists():
        return FALLBACK_TEACHER_ROSTER_ACCOUNTS

    try:
        with TEACHER_ROSTER_TEMPLATE_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            accounts: dict[str, str] = {}
            for row in reader:
                email = (row.get("email") or "").strip().lower()
                name = (row.get("name") or "").strip()
                if not email or not name:
                    continue
                accounts[email] = name

            return list(accounts.items()) or FALLBACK_TEACHER_ROSTER_ACCOUNTS
    except OSError:
        return FALLBACK_TEACHER_ROSTER_ACCOUNTS


def get_teacher_roster_access_map() -> dict[str, str]:
    return {
        email.strip().lower(): full_name
        for email, full_name in load_teacher_roster_template_accounts()
    }


def is_teacher_roster_access_account(email: str) -> bool:
    return (email or "").strip().lower() in get_teacher_roster_access_map()

def normalized_code(name: str, course_type: str, raw_code: str, used_codes: set[str]) -> str:
    base = (raw_code or "").strip().rstrip(".")
    if not base:
        name_token = re.sub(r"[^A-Za-z0-9]+", "", name).upper()[:10] or "COURSE"
        type_token = re.sub(r"[^A-Za-z0-9]+", "", course_type).upper()[:4] or "TYPE"
        base = f"{name_token}-{type_token}"

    candidate = base
    suffix = 2
    while candidate in used_codes:
        type_token = re.sub(r"[^A-Za-z0-9]+", "", course_type).upper()[:4] or "TYPE"
        candidate = f"{base}-{type_token}-{suffix}"
        suffix += 1

    used_codes.add(candidate)
    return candidate

def seed_default_courses_if_empty():
    semester = current_semester_label()
    with Session(engine) as db:
        used_codes = {course.code for course in db.query(models.Course).all()}
        existing_pairs = {
            (course.name, course.course_type, course.semester)
            for course in db.query(models.Course).all()
        }

        for name, course_type, raw_code in DEFAULT_COURSES:
            pair = (name, course_type, semester)
            if pair in existing_pairs:
                continue

            course = models.Course(
                id=uuid.uuid4(),
                name=name,
                course_type=course_type,
                code=normalized_code(name, course_type, raw_code, used_codes),
                semester=semester,
                lecturer_id=None,
            )
            db.add(course)
            existing_pairs.add(pair)

        db.commit()

def seed_default_lecturers():
    with Session(engine) as db:
        for full_name, raw_email, title in DEFAULT_LECTURERS:
            email = raw_email.strip().lower()

            user = db.query(models.User).filter(models.User.email == email).first()
            if not user:
                user = models.User(
                    id=uuid.uuid4(),
                    email=email,
                    name=full_name,
                    role=models.UserRole.TEACHER,
                )
                db.add(user)
            else:
                user.role = models.UserRole.TEACHER
                user.name = full_name

            profile = db.query(models.LecturerProfile).filter(models.LecturerProfile.email == email).first()
            now = datetime.now(timezone.utc)
            if not profile:
                profile = models.LecturerProfile(
                    id=uuid.uuid4(),
                    email=email,
                    title=title,
                    full_name=full_name,
                    must_change_password=False,
                    created_at=now,
                    updated_at=now,
                )
                db.add(profile)
            else:
                profile.title = title
                profile.full_name = full_name
                profile.updated_at = now

            ensure_local_credential(
                db,
                username=email,
                password=DEFAULT_LECTURER_PASSWORD,
                role="TEACHER",
                user_email=email,
            )

        db.commit()


def seed_teacher_roster_access_accounts():
    with Session(engine) as db:
        for email, full_name in load_teacher_roster_template_accounts():
            normalized_email = email.strip().lower()
            user = db.query(models.User).filter(models.User.email == normalized_email).first()
            if not user:
                user = models.User(
                    id=uuid.uuid4(),
                    email=normalized_email,
                    name=full_name,
                    role=models.UserRole.TEACHER,
                    is_admin=False,
                )
                db.add(user)
            else:
                user.name = full_name
                user.role = models.UserRole.TEACHER

            profile = db.query(models.LecturerProfile).filter(models.LecturerProfile.email == normalized_email).first()
            now = datetime.now(timezone.utc)
            if not profile:
                profile = models.LecturerProfile(
                    id=uuid.uuid4(),
                    email=normalized_email,
                    title="Lecturer",
                    full_name=full_name,
                    must_change_password=False,
                    created_at=now,
                    updated_at=now,
                )
                db.add(profile)
            else:
                profile.title = profile.title or "Lecturer"
                profile.full_name = full_name
                profile.must_change_password = False
                profile.updated_at = now

            ensure_local_credential(
                db,
                username=normalized_email,
                password=DEFAULT_LECTURER_PASSWORD,
                role="TEACHER",
                user_email=normalized_email,
            )

        db.commit()

def hash_local_password(password: str) -> str:
    # DB needs credential records; store hashes instead of plain text.
    digest = hashlib.sha256(password.encode("utf-8")).hexdigest()
    return f"sha256${digest}"

def ensure_local_credential(
    db: Session,
    username: str,
    password: str,
    role: str,
    user_email: str | None,
    force_update: bool = False,
):
    now = datetime.now(timezone.utc)
    normalized_username = username.strip().lower()
    normalized_role = role.strip().upper()

    existing = (
        db.query(models.LocalCredential)
        .filter(
            models.LocalCredential.username == normalized_username,
            models.LocalCredential.role == normalized_role,
        )
        .first()
    )

    if existing:
        existing.user_email = user_email
        existing.active = True
        if force_update:
            existing.password_hash = hash_local_password(password)
        existing.updated_at = now
        return existing

    record = models.LocalCredential(
        id=uuid.uuid4(),
        username=normalized_username,
        password_hash=hash_local_password(password),
        role=normalized_role,
        user_email=user_email,
        active=True,
        created_at=now,
        updated_at=now,
    )
    db.add(record)
    return record

def seed_superuser_credential():
    with Session(engine) as db:
        ensure_local_credential(
            db,
            username=SUPERUSER_USERNAME,
            password=SUPERUSER_PASSWORD,
            role="SUPERUSER",
            user_email=None,
        )
        db.commit()


def seed_superuser_account():
    with Session(engine) as db:
        email = SUPERUSER_EMAIL.strip().lower()
        user = db.query(models.User).filter(models.User.email == email).first()
        if not user:
            user = models.User(
                id=uuid.uuid4(),
                email=email,
                name=SUPERUSER_DISPLAY_NAME,
                role=models.UserRole.TEACHER,
                is_admin=True,
            )
            db.add(user)
        else:
            user.name = SUPERUSER_DISPLAY_NAME
            user.role = models.UserRole.TEACHER
            user.is_admin = True

        profile = db.query(models.LecturerProfile).filter(models.LecturerProfile.email == email).first()
        now = datetime.now(timezone.utc)
        if not profile:
            profile = models.LecturerProfile(
                id=uuid.uuid4(),
                email=email,
                title="Superuser",
                full_name=SUPERUSER_DISPLAY_NAME,
                must_change_password=False,
                created_at=now,
                updated_at=now,
            )
            db.add(profile)
        else:
            profile.title = "Superuser"
            profile.full_name = SUPERUSER_DISPLAY_NAME
            profile.must_change_password = False
            profile.updated_at = now

        ensure_local_credential(
            db,
            username=email,
            password=SUPERUSER_PASSWORD,
            role="TEACHER",
            user_email=email,
        )
        ensure_local_credential(
            db,
            username=email,
            password=SUPERUSER_PASSWORD,
            role="SUPERUSER",
            user_email=email,
        )
        db.commit()


def seed_mental_access_students():
    with Session(engine) as db:
        for email, full_name in MENTAL_ACCESS_STUDENTS:
            normalized_email = email.strip().lower()
            user = db.query(models.User).filter(models.User.email == normalized_email).first()
            if not user:
                user = models.User(
                    id=uuid.uuid4(),
                    email=normalized_email,
                    name=full_name,
                    role=models.UserRole.STUDENT,
                    is_admin=False,
                )
                db.add(user)
            else:
                user.name = full_name
                user.role = models.UserRole.STUDENT

        db.commit()

ensure_schema_upgrades()
seed_default_courses_if_empty()
seed_default_lecturers()
seed_teacher_roster_access_accounts()
seed_superuser_credential()
seed_superuser_account()

SUPERUSER_KEY = os.getenv("SUPERUSER_KEY", "admin123")
TOKEN_WINDOW_SECONDS = 10
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
SUPABASE_URL = os.getenv("SUPABASE_URL", "")

def to_utc_aware(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)

def verify_admin(x_admin_key: str = Header(default="")):
    if x_admin_key != SUPERUSER_KEY:
        raise HTTPException(status_code=401, detail="Invalid admin key")

def verify_admin_access(
    x_admin_key: str = Header(default=""),
    credentials: HTTPAuthorizationCredentials | None = Depends(optional_security),
    db: Session = Depends(get_db),
):
    if x_admin_key == SUPERUSER_KEY:
        return {"mode": "key", "email": None}

    if not credentials:
        raise HTTPException(status_code=401, detail="Admin authentication required")

    payload = auth.verify_token(credentials.credentials)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid token")

    email = (payload.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(status_code=401, detail="Authenticated user has no email")

    user_metadata = payload.get("user_metadata") or {}
    display_name = (
        user_metadata.get("name")
        or user_metadata.get("full_name")
        or email.split("@")[0]
    )

    user = db.query(models.User).filter(models.User.email == email).first()
    if not user:
        user = models.User(
            id=uuid.uuid4(),
            email=email,
            name=display_name,
            role=models.UserRole.TEACHER,
            is_admin=False,
        )
        db.add(user)
        db.commit()
        db.refresh(user)
    elif not user.name:
        user.name = display_name
        db.commit()

    has_seeded_admin = db.query(models.User).filter(models.User.is_admin == True).first() is not None
    bootstrap_allowed = (not has_seeded_admin) and user.role == models.UserRole.TEACHER

    if not user.is_admin and not bootstrap_allowed:
        raise HTTPException(status_code=403, detail="Admin access denied")

    return {
        "mode": "token",
        "email": email,
        "isBootstrap": bootstrap_allowed,
        "isAdmin": bool(user.is_admin),
    }

def parse_uuid_or_400(value: str, field_name: str):
    try:
        return uuid.UUID(value)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid {field_name}")

def get_latest_emergency_message(db: Session) -> str | None:
    latest = (
        db.query(models.EmergencyBroadcast)
        .filter(models.EmergencyBroadcast.active == True)
        .order_by(models.EmergencyBroadcast.created_at.desc())
        .first()
    )
    if not latest:
        return None
    return latest.message

def generate_temp_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))

def _supabase_admin_headers() -> dict[str, str] | None:
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        return None
    return {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
    }

def _find_supabase_user_id_by_email(email: str) -> str | None:
    headers = _supabase_admin_headers()
    if not headers:
        return None

    try:
        res = requests.get(
            f"{SUPABASE_URL}/auth/v1/admin/users",
            headers=headers,
            params={"page": 1, "per_page": 1000},
            timeout=10,
        )
        if not res.ok:
            return None

        payload = res.json()
        users = payload.get("users", []) if isinstance(payload, dict) else []
        for item in users:
            if (item.get("email") or "").lower() == email.lower():
                return item.get("id")
    except requests.RequestException:
        return None

    return None

def upsert_supabase_auth_user(
    email: str,
    password: str,
    user_metadata: dict[str, object] | None = None,
) -> tuple[bool, str]:
    headers = _supabase_admin_headers()
    if not headers:
        return False, "SUPABASE_SERVICE_ROLE_KEY is not configured."

    body = {
        "email": email,
        "password": password,
        "email_confirm": True,
        "user_metadata": user_metadata or {},
    }

    try:
        create_res = requests.post(
            f"{SUPABASE_URL}/auth/v1/admin/users",
            headers=headers,
            json=body,
            timeout=10,
        )
        if create_res.ok:
            return True, "Auth account provisioned."

        user_id = _find_supabase_user_id_by_email(email)
        if not user_id:
            return False, "Unable to provision Supabase auth user."

        update_res = requests.put(
            f"{SUPABASE_URL}/auth/v1/admin/users/{user_id}",
            headers=headers,
            json={
                "password": password,
                "email_confirm": True,
                "user_metadata": user_metadata or {},
            },
            timeout=10,
        )
        if not update_res.ok:
            return False, "Unable to update Supabase auth account password."

        return True, "Auth account password refreshed."
    except requests.RequestException:
        return False, "Supabase auth provisioning request failed."


def ensure_supabase_auth_user(
    email: str,
    password: str,
    user_metadata: dict[str, object] | None = None,
) -> tuple[bool, str]:
    if not _supabase_admin_headers():
        return False, "SUPABASE_SERVICE_ROLE_KEY is not configured."

    if _find_supabase_user_id_by_email(email):
        return True, "Auth account already present."

    return upsert_supabase_auth_user(email=email, password=password, user_metadata=user_metadata)


def upsert_supabase_lecturer(email: str, full_name: str, title: str, password: str) -> tuple[bool, str]:
    return upsert_supabase_auth_user(
        email=email,
        password=password,
        user_metadata={
            "name": full_name,
            "title": title,
        },
    )


def upsert_supabase_student(email: str, full_name: str, password: str) -> tuple[bool, str]:
    return upsert_supabase_auth_user(
        email=email,
        password=password,
        user_metadata={
            "name": full_name,
            "mentalAccess": True,
        },
    )


def provision_mental_access_student_auth():
    if not _supabase_admin_headers():
        return

    for email, full_name in MENTAL_ACCESS_STUDENTS:
        upsert_supabase_student(email, full_name, DEFAULT_TEST_STUDENT_PASSWORD)


def provision_teacher_roster_auth():
    if not _supabase_admin_headers():
        return

    for email, full_name in load_teacher_roster_template_accounts():
        # Upsert (not just ensure) so stale existing auth users also get the expected default password.
        upsert_supabase_lecturer(
            email=email,
            full_name=full_name,
            title="Lecturer",
            password=DEFAULT_LECTURER_PASSWORD,
        )


def provision_superuser_auth():
    if not _supabase_admin_headers():
        return

    ensure_supabase_auth_user(
        email=SUPERUSER_EMAIL,
        password=SUPERUSER_PASSWORD,
        user_metadata={
            "name": SUPERUSER_DISPLAY_NAME,
            "title": "Superuser",
            "isAdmin": True,
        },
    )


seed_mental_access_students()
provision_mental_access_student_auth()
provision_teacher_roster_auth()
provision_superuser_auth()

def ensure_teacher_can_manage_course(db: Session, teacher_email: str, course_uuid: uuid.UUID):
    if is_teacher_roster_access_account(teacher_email):
        return

    assignment = (
        db.query(models.TeacherCourseAssignment)
        .filter(
            models.TeacherCourseAssignment.teacher_email == teacher_email,
            models.TeacherCourseAssignment.course_id == course_uuid,
        )
        .first()
    )
    if assignment:
        return

    course = db.query(models.Course).filter(models.Course.id == course_uuid).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")

    # Legacy fallback: allow management if lecturer_id points to the same teacher user.
    teacher_user = db.query(models.User).filter(models.User.email == teacher_email).first()
    if teacher_user and course.lecturer_id == teacher_user.id:
        return

    raise HTTPException(status_code=403, detail="You are not assigned to manage this course")


def _is_valid_roster_email(email: str) -> bool:
    # Testing-mode validation focuses on email presence/shape rather than specific domain allowlists.
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email))


def _finalize_roster_entries(entries: list[tuple[str, str]], invalid_rows: int) -> tuple[list[tuple[str, str]], int]:
    dedup: dict[str, str] = {}
    for email, name in entries:
        normalized_email = (email or "").strip().lower()
        normalized_name = (name or "").strip()

        if not normalized_email or not _is_valid_roster_email(normalized_email):
            invalid_rows += 1
            continue

        dedup[normalized_email] = normalized_name

    return [(email, dedup[email]) for email in dedup], invalid_rows


def parse_roster_csv(file_bytes: bytes) -> tuple[list[tuple[str, str]], int]:
    decoded = file_bytes.decode("utf-8-sig", errors="ignore")
    stream = io.StringIO(decoded)

    entries: list[tuple[str, str]] = []
    invalid_rows = 0

    first_line = stream.readline()
    stream.seek(0)
    has_header = "email" in first_line.lower()

    if has_header:
        reader = csv.DictReader(stream)
        for row in reader:
            email = (row.get("email") or row.get("student_email") or "").strip().lower()
            name = (row.get("name") or row.get("student_name") or "").strip()
            if not email:
                invalid_rows += 1
                continue
            entries.append((email, name))
    else:
        reader = csv.reader(stream)
        for row in reader:
            if not row:
                continue
            email = (row[0] or "").strip().lower()
            name = (row[1] if len(row) > 1 else "").strip()
            if not email:
                invalid_rows += 1
                continue
            entries.append((email, name))

    return _finalize_roster_entries(entries, invalid_rows)


def parse_roster_xlsx(file_bytes: bytes) -> tuple[list[tuple[str, str]], int]:
    invalid_rows = 0
    entries: list[tuple[str, str]] = []

    workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    first_row = next(rows, None)
    if first_row is None:
        return [], 0

    first_values = [str(item).strip() if item is not None else "" for item in first_row]
    lowered = [item.lower() for item in first_values]

    has_header = any("email" in item for item in lowered)

    def row_to_email_name(row_values: tuple) -> tuple[str, str]:
        email = (str(row_values[0]).strip().lower() if len(row_values) > 0 and row_values[0] is not None else "")
        name = (str(row_values[1]).strip() if len(row_values) > 1 and row_values[1] is not None else "")
        return email, name

    if has_header:
        email_index = next((idx for idx, item in enumerate(lowered) if item in {"email", "student_email"}), None)
        if email_index is None:
            return [], 1

        name_index = next((idx for idx, item in enumerate(lowered) if item in {"name", "student_name"}), None)

        for row in rows:
            if row is None:
                continue
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            email = ""
            name = ""
            if email_index < len(row) and row[email_index] is not None:
                email = str(row[email_index]).strip().lower()
            if name_index is not None and name_index < len(row) and row[name_index] is not None:
                name = str(row[name_index]).strip()

            if not email:
                invalid_rows += 1
                continue
            entries.append((email, name))
    else:
        email, name = row_to_email_name(first_row)
        if email:
            entries.append((email, name))
        elif any(item is not None and str(item).strip() for item in first_row):
            invalid_rows += 1

        for row in rows:
            if row is None:
                continue
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            email, name = row_to_email_name(row)
            if not email:
                invalid_rows += 1
                continue
            entries.append((email, name))

    return _finalize_roster_entries(entries, invalid_rows)


def parse_roster_file(file_bytes: bytes, filename: str | None) -> tuple[list[tuple[str, str]], int]:
    extension = Path(filename or "").suffix.lower()

    if extension == ".csv":
        return parse_roster_csv(file_bytes)
    if extension in {".xlsx", ".xlsm"}:
        return parse_roster_xlsx(file_bytes)

    raise HTTPException(status_code=400, detail="Unsupported roster file type. Use .csv or .xlsx")

app = FastAPI(title="Smart Multi-Modal Attendance API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def read_root():
    return {"message": "Smart Multi-Modal Attendance API is running!"}

@app.post("/auth/login", response_model=schemas.LoginResponse)
def login(request: schemas.LoginRequest, db: Session = Depends(get_db)):
    payload = auth.verify_token(request.providerToken)

    if not payload:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    try:
        role = UserRole(request.role)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid role")

    email = payload.get("email")
    if not email:
        raise HTTPException(status_code=401, detail="Email missing in token")
    
    user = db.query(models.User).filter(models.User.email == email).first()

    if not user:
        user = models.User(
            email=email,
            name=payload.get("user_metadata", {}).get("name"),
            role=role
        )
        db.add(user)

    elif user.role != role:
        user.role = request.role

    db.commit()
    db.refresh(user)

    return {
        "userId": str(user.id),
        "email": user.email,
        "name": user.name,
        "role": user.role.value
    }

@app.post("/auth/session-sync", response_model=schemas.AuthSessionSyncResponse)
def auth_session_sync(
    request: schemas.AuthSessionSyncRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    email = (current_user.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(status_code=401, detail="Authenticated user email is missing")

    name = (current_user.get("name") or "").strip() or email.split("@")[0]
    mental_access_name = get_mental_access_student_name(email)

    try:
        requested_role = UserRole(request.role.strip().upper())
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid role. Use STUDENT or TEACHER.")

    if requested_role == UserRole.STUDENT and mental_access_name:
        name = mental_access_name

    user = db.query(models.User).filter(models.User.email == email).first()

    if not user:
        user = models.User(
            id=uuid.uuid4(),
            email=email,
            name=name,
            role=requested_role,
            is_admin=False,
        )
        db.add(user)
    else:
        user.name = name
        if requested_role == UserRole.TEACHER:
            user.role = UserRole.TEACHER
        elif user.role != UserRole.TEACHER:
            user.role = UserRole.STUDENT

    db.commit()
    db.refresh(user)

    return {
        "userId": str(user.id),
        "email": user.email,
        "name": user.name,
        "role": user.role.value,
        "isAdmin": bool(user.is_admin),
    }

@app.post("/sessions", response_model=schemas.SessionResponse)
def create_session(session_req: schemas.SessionCreate, db: Session = Depends(get_db)):
    course = db.query(models.Course).filter(models.Course.id == session_req.course_id).first()

    if not course:
        raise HTTPException(status_code=404, detail="Course not found")

    session = models.Session(
        id=uuid.uuid4(),
        course_id=session_req.course_id,
        lecturer_id=course.lecturer_id, 
        active=True,
        created_at=datetime.now(timezone.utc),
        expires_at=datetime.now(timezone.utc) + timedelta(hours=1)
    )

    # A new class session clears prior emergency protocol banner state.
    db.query(models.EmergencyBroadcast).filter(
        models.EmergencyBroadcast.active == True
    ).update({"active": False}, synchronize_session=False)

    db.add(session)
    db.commit()
    db.refresh(session)

    return {
        "sessionId": session.id,
        "isActive": session.active,
        "createdAt": session.created_at
    }

@app.get("/courses/teacher")
def get_teacher_courses(semester: str | None = None, teacherEmail: str | None = None, db: Session = Depends(get_db)):
    query = db.query(models.Course)
    if semester:
        query = query.filter(models.Course.semester == semester)

    if teacherEmail:
        normalized_teacher_email = teacherEmail.strip().lower()
        if not is_teacher_roster_access_account(normalized_teacher_email):
            assigned_ids = {
                assignment.course_id
                for assignment in db.query(models.TeacherCourseAssignment)
                .filter(models.TeacherCourseAssignment.teacher_email == normalized_teacher_email)
                .all()
            }
            if assigned_ids:
                query = query.filter(models.Course.id.in_(assigned_ids))
            else:
                return []

    courses = query.order_by(models.Course.name.asc()).all()

    active_sessions = {
        str(session.course_id): str(session.id)
        for session in db.query(models.Session).filter(models.Session.active == True).all()
    }

    return [
        {
            "id": str(course.id),
            "name": course.name,
            "courseType": course.course_type,
            "code": course.code,
            "semester": course.semester,
            "activeSessionId": active_sessions.get(str(course.id))
        }
        for course in courses
    ]

@app.get("/courses/catalog")
def get_course_catalog(semester: str | None = None, db: Session = Depends(get_db)):
    query = db.query(models.Course)
    if semester:
        query = query.filter(models.Course.semester == semester)

    courses = query.order_by(models.Course.name.asc()).all()

    return [
        {
            "id": str(course.id),
            "name": course.name,
            "courseType": course.course_type,
            "code": course.code,
            "semester": course.semester,
        }
        for course in courses
    ]

@app.post("/teacher/courses/assign")
def assign_course_to_teacher(req: schemas.TeacherCourseAssignRequest, db: Session = Depends(get_db)):
    course = db.query(models.Course).filter(models.Course.id == req.course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")

    existing = (
        db.query(models.TeacherCourseAssignment)
        .filter(
            models.TeacherCourseAssignment.teacher_email == req.teacher_email,
            models.TeacherCourseAssignment.course_id == req.course_id,
        )
        .first()
    )

    if existing:
        return {"success": True, "message": "Course already assigned."}

    assignment = models.TeacherCourseAssignment(
        id=uuid.uuid4(),
        teacher_email=req.teacher_email,
        course_id=req.course_id,
    )
    db.add(assignment)
    db.commit()

    return {"success": True, "message": "Course assigned to teacher."}

@app.delete("/teacher/courses/{courseId}/assign")
def unassign_course_from_teacher(courseId: str, teacherEmail: str, db: Session = Depends(get_db)):
    course_uuid = parse_uuid_or_400(courseId, "courseId")

    assignment = (
        db.query(models.TeacherCourseAssignment)
        .filter(
            models.TeacherCourseAssignment.teacher_email == teacherEmail,
            models.TeacherCourseAssignment.course_id == course_uuid,
        )
        .first()
    )
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")

    db.delete(assignment)
    db.commit()

    return {"success": True, "message": "Course removed from your list."}

@app.post("/courses/requests")
def create_course_request(req: schemas.CourseRequestCreate, db: Session = Depends(get_db)):
    new_request = models.CourseRequest(
        id=uuid.uuid4(),
        teacher_email=req.teacher_email,
        name=req.name.strip(),
        course_type=req.course_type.strip() or "L",
        code=(req.code or "").strip() or None,
        semester=req.semester.strip() or current_semester_label(),
        status="PENDING",
    )
    db.add(new_request)
    db.commit()
    db.refresh(new_request)

    return {
        "id": str(new_request.id),
        "status": new_request.status,
        "message": "Request submitted for DB admin approval.",
    }

@app.get("/teacher/courses/{courseId}/students")
def get_course_student_registry(
    courseId: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    course_uuid = parse_uuid_or_400(courseId, "courseId")
    teacher_email = (current_user.get("email") or "").strip().lower()
    ensure_teacher_can_manage_course(db, teacher_email, course_uuid)

    roster_map = get_effective_course_student_roster_map(db, course_uuid)
    students = [
        {
            "email": email,
            "name": name or None,
        }
        for email, name in sorted(roster_map.items())
    ]

    return {
        "courseId": courseId,
        "count": len(students),
        "students": students,
    }

@app.post("/teacher/courses/{courseId}/students/upload", response_model=schemas.CourseRosterUploadSummary)
async def upload_course_student_registry(
    courseId: str,
    rosterFile: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    course_uuid = parse_uuid_or_400(courseId, "courseId")
    teacher_email = (current_user.get("email") or "").strip().lower()
    ensure_teacher_can_manage_course(db, teacher_email, course_uuid)

    file_bytes = await rosterFile.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded roster file is empty")

    parsed_entries, invalid_rows = parse_roster_file(file_bytes, rosterFile.filename)
    if not parsed_entries:
        raise HTTPException(status_code=400, detail="No valid student rows found in roster file")

    db.query(models.CourseStudentRegistry).filter(
        models.CourseStudentRegistry.course_id == course_uuid
    ).delete(synchronize_session=False)

    now = datetime.now(timezone.utc)
    for email, name in parsed_entries:
        db.add(
            models.CourseStudentRegistry(
                id=uuid.uuid4(),
                course_id=course_uuid,
                student_email=email,
                student_name=name or None,
                uploaded_by=teacher_email,
                created_at=now,
            )
        )

    db.commit()

    return {
        "courseId": courseId,
        "uploadedCount": len(parsed_entries),
        "invalidRows": invalid_rows,
        "message": "Student registry uploaded and activated for this course. Mental-access students remain validated by default.",
    }

@app.get("/semesters")
def get_semesters(db: Session = Depends(get_db)):
    semesters = sorted({
        course.semester for course in db.query(models.Course).all() if course.semester
    })

    if not semesters:
        semesters = [current_semester_label()]

    return {"semesters": semesters}

@app.post("/courses")
def create_course(course_req: schemas.CourseCreate, db: Session = Depends(get_db)):
    existing = db.query(models.Course).filter(models.Course.code == course_req.code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Course code already exists")

    new_course = models.Course(
        id=uuid.uuid4(),
        name=course_req.name.strip(),
        course_type=course_req.course_type.strip() or "L",
        code=course_req.code.strip(),
        semester=course_req.semester.strip() or current_semester_label(),
        lecturer_id=None,
    )

    db.add(new_course)
    db.commit()
    db.refresh(new_course)

    return {
        "id": str(new_course.id),
        "name": new_course.name,
        "courseType": new_course.course_type,
        "code": new_course.code,
        "semester": new_course.semester,
    }

@app.delete("/courses/{courseId}")
def delete_course(courseId: str, db: Session = Depends(get_db)):
    course_uuid = parse_uuid_or_400(courseId, "courseId")
    course = db.query(models.Course).filter(models.Course.id == course_uuid).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")

    sessions = db.query(models.Session).filter(models.Session.course_id == course.id).all()
    session_ids = [session.id for session in sessions]

    if session_ids:
        db.query(models.AttendanceLog).filter(models.AttendanceLog.session_id.in_(session_ids)).delete(synchronize_session=False)
        db.query(models.Session).filter(models.Session.id.in_(session_ids)).delete(synchronize_session=False)

    db.query(models.TeacherCourseAssignment).filter(
        models.TeacherCourseAssignment.course_id == course_uuid
    ).delete(synchronize_session=False)

    db.delete(course)
    db.commit()

    return {"success": True, "message": "Course removed successfully."}

@app.get("/courses/student")
def get_student_courses(db: Session = Depends(get_db)):
    courses = db.query(models.Course).all()

    active_course_ids = {
        session.course_id
        for session in db.query(models.Session).filter(models.Session.active == True).all()
    }

    return [
        {
            "id": str(course.id),
            "name": course.name,
            "courseType": course.course_type,
            "code": course.code,
            "active": course.id in active_course_ids
        }
        for course in courses
    ]

@app.get("/sessions/active")
def get_active_session(courseId: str, db: Session = Depends(get_db)):
    try:
        course_uuid = uuid.UUID(courseId)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid courseId")

    session = (
        db.query(models.Session)
        .filter(
            models.Session.course_id == course_uuid,
            models.Session.active == True
        )
        .order_by(models.Session.created_at.desc())
        .first()
    )

    if not session:
        return {
            "active": False,
            "sessionId": None
        }

    return {
        "active": True,
        "sessionId": str(session.id)
    }

@app.post("/sessions/{sessionId}/end")
def end_session(sessionId: str, db: Session = Depends(get_db)):
    session_uuid = parse_uuid_or_400(sessionId, "sessionId")
    session = db.query(models.Session).filter(models.Session.id == session_uuid).first()

    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    if not session.active:
        return {
            "sessionId": str(session.id),
            "isActive": False,
            "message": "Session already ended."
        }

    session.active = False
    session.expires_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(session)

    return {
        "sessionId": str(session.id),
        "isActive": session.active,
        "message": "Session ended successfully."
    }

@app.get("/qr/{sessionId}")
def get_qr_token(sessionId: str, db: Session = Depends(get_db)):
    session_uuid = parse_uuid_or_400(sessionId, "sessionId")
    session = db.query(models.Session).filter(models.Session.id == session_uuid).first()

    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    if not session.active:
        emergency_message = get_latest_emergency_message(db)
        if emergency_message:
            raise HTTPException(status_code=400, detail=f"Emergency protocol: {emergency_message}")
        raise HTTPException(status_code=400, detail="Session is not active")

    if session.expires_at and to_utc_aware(session.expires_at) < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Session expired")

    window = int(time.time()) // TOKEN_WINDOW_SECONDS

    raw = f"{sessionId}:{window}"
    token = hashlib.sha256(raw.encode()).hexdigest()

    return {
        "token": token,
        "expiresAt": datetime.now(timezone.utc).isoformat()
    }

@app.post("/attendance", response_model=schemas.ScanResponse)
def submit_scan(
    scan_req: schemas.ScanRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    client_ip = request.client.host

    student_email = (current_user.get("email") or "").strip().lower()
    if not re.match(r"^[^@]+@[^@]+\.elte\.hu$", student_email):
        raise HTTPException(
            status_code=403,
            detail="Student login must use ELTE Microsoft SSO email (*.elte.hu).",
        )

    student_name = get_mental_access_student_name(student_email) or (current_user.get("name") or "").strip() or student_email.split("@")[0]

    student = db.query(models.User).filter(models.User.email == student_email).first()
    if not student:
        student = models.User(
            id=uuid.uuid4(),
            email=student_email,
            name=student_name,
            role=models.UserRole.STUDENT,
        )
        db.add(student)
        db.flush()
    elif student.role != models.UserRole.STUDENT:
        student.role = models.UserRole.STUDENT

    # Name comes from Microsoft SSO for now; registration-list validation can be layered later.
    student.name = student_name

    sessions = db.query(models.Session).filter(models.Session.active == True).all()

    if not sessions:
        emergency_message = get_latest_emergency_message(db)
        if emergency_message:
            raise HTTPException(status_code=404, detail=f"Emergency protocol: {emergency_message}")
        raise HTTPException(status_code=404, detail="No active sessions")

    current_window = int(time.time()) // TOKEN_WINDOW_SECONDS

    valid_session = None

    for session in sessions:
        raw = f"{session.id}:{current_window}"
        expected_token = hashlib.sha256(raw.encode()).hexdigest()

        prev_raw = f"{session.id}:{current_window - 1}"
        prev_token = hashlib.sha256(prev_raw.encode()).hexdigest()

        if scan_req.token in [expected_token, prev_token]:
            valid_session = session
            break

    if not valid_session:
        raise HTTPException(status_code=401, detail="Invalid or expired QR token")

    uploaded_roster_map = get_uploaded_course_student_roster_map(db, valid_session.course_id)

    if uploaded_roster_map:
        if student_email not in uploaded_roster_map and not is_mental_access_student(student_email):
            raise HTTPException(
                status_code=403,
                detail="Student is not registered in the uploaded course roster.",
            )

        uploaded_name = (uploaded_roster_map.get(student_email) or "").strip()
        if uploaded_name:
            student.name = uploaded_name
    elif is_mental_access_student(student_email):
        student.name = get_mental_access_student_name(student_email) or student.name

    existing = db.query(models.AttendanceLog).filter(
        models.AttendanceLog.session_id == valid_session.id,
        models.AttendanceLog.device_id == scan_req.deviceToken
    ).first()

    if existing:
        raise HTTPException(status_code=400, detail="Device already scanned")

    log = models.AttendanceLog(
        session_id=valid_session.id,
        user_id=student.id,
        device_id=scan_req.deviceToken,
        ip_address=client_ip,
        qr_timestamp=datetime.now(timezone.utc),
        status="PRESENT"
    )

    db.add(log)
    db.commit()
    db.refresh(log)

    return {
        "success": True,
        "message": "Attendance recorded successfully."
    }

@app.get("/sessions/{sessionId}/export")
def export_attendance(sessionId: str, db: Session = Depends(get_db)):
    session_uuid = parse_uuid_or_400(sessionId, "sessionId")
    session = db.query(models.Session).filter(models.Session.id == session_uuid).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    # LEFT JOIN (IMPORTANT FIX)
    rows = (
        db.query(models.AttendanceLog, models.User)
        .outerjoin(models.User, models.AttendanceLog.user_id == models.User.id)
        .filter(models.AttendanceLog.session_id == session_uuid)
        .all()
    )

    output = io.StringIO()
    writer = csv.writer(output)

    # headers = separate columns
    writer.writerow([
        "User ID",
        "User name",
        "Device ID",
        "IP Address",
        "Timestamp",
        "Status"
    ])

    for log, user in rows:
        user_id_value = ""
        if user and user.email:
            user_id_value = user.email.split("@")[0]

        writer.writerow([
            user_id_value,
            user.name if user else "",
            log.device_id,
            log.ip_address,
            log.timestamp,
            log.status
        ])

    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=attendance_{sessionId}.csv"
        }
    )

@app.get("/admin/overview", response_model=schemas.AdminOverview)
def admin_overview(_: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    return {
        "users": db.query(models.User).count(),
        "courses": db.query(models.Course).count(),
        "activeSessions": db.query(models.Session).filter(models.Session.active == True).count(),
        "totalSessions": db.query(models.Session).count(),
        "attendanceLogs": db.query(models.AttendanceLog).count(),
    }

@app.get("/admin/sessions/active")
def admin_active_sessions(_: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    sessions = (
        db.query(models.Session, models.Course)
        .join(models.Course, models.Session.course_id == models.Course.id)
        .filter(models.Session.active == True)
        .order_by(models.Session.created_at.desc())
        .all()
    )

    return [
        {
            "sessionId": str(session.id),
            "courseId": str(course.id),
            "courseName": course.name,
            "courseCode": course.code,
            "semester": course.semester,
            "startedAt": session.created_at,
        }
        for session, course in sessions
    ]

@app.post("/admin/sessions/end-all")
def admin_end_all_sessions(req: schemas.AdminEndAllSessionsRequest, _: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    comment = req.comment.strip()
    if not comment:
        raise HTTPException(status_code=400, detail="Comment is required for emergency protocol.")

    now = datetime.now(timezone.utc)
    active_sessions = db.query(models.Session).filter(models.Session.active == True).all()

    ended_count = 0
    for session in active_sessions:
        session.active = False
        session.expires_at = now
        ended_count += 1

    db.query(models.EmergencyBroadcast).filter(
        models.EmergencyBroadcast.active == True
    ).update({"active": False}, synchronize_session=False)

    db.add(
        models.EmergencyBroadcast(
            id=uuid.uuid4(),
            message=comment,
            active=True,
            created_at=now,
        )
    )

    db.commit()

    return {
        "success": True,
        "endedSessions": ended_count,
        "message": "Emergency protocol executed. All active sessions ended.",
        "comment": comment,
    }

@app.get("/system/emergency-message")
def get_emergency_message(db: Session = Depends(get_db)):
    message = get_latest_emergency_message(db)
    return {
        "active": bool(message),
        "message": message,
    }

@app.get("/admin/course-requests")
def admin_course_requests(_: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    requests = (
        db.query(models.CourseRequest)
        .order_by(models.CourseRequest.created_at.desc())
        .all()
    )

    return [
        {
            "id": str(item.id),
            "teacherEmail": item.teacher_email,
            "name": item.name,
            "courseType": item.course_type,
            "code": item.code,
            "semester": item.semester,
            "status": item.status,
            "createdAt": item.created_at,
            "reviewedAt": item.reviewed_at,
            "reviewerNote": item.reviewer_note,
        }
        for item in requests
    ]

@app.get("/admin/lecturers")
def admin_list_lecturers(_: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    profiles_by_email = {
        item.email.lower(): item
        for item in db.query(models.LecturerProfile).all()
    }

    lecturers = (
        db.query(models.User)
        .filter(models.User.role == models.UserRole.TEACHER)
        .order_by(models.User.created_at.desc())
        .all()
    )

    return [
        {
            "id": str(item.id),
            "email": item.email,
            "name": item.name,
            "title": (profiles_by_email.get(item.email.lower()).title if profiles_by_email.get(item.email.lower()) else "Lecturer"),
            "fullName": (profiles_by_email.get(item.email.lower()).full_name if profiles_by_email.get(item.email.lower()) else item.name),
            "mustChangePassword": bool(profiles_by_email.get(item.email.lower()).must_change_password) if profiles_by_email.get(item.email.lower()) else False,
            "role": item.role.value,
            "isAdmin": bool(item.is_admin),
            "createdAt": item.created_at,
        }
        for item in lecturers
    ]

@app.post("/admin/lecturers")
def admin_create_lecturer(req: schemas.AdminLecturerCreateRequest, _: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    email = req.email.strip().lower()
    title = req.title.strip() or "Lecturer"
    full_name = req.full_name.strip()
    is_admin = bool(req.is_admin)

    if not email:
        raise HTTPException(status_code=400, detail="Lecturer email is required.")
    if not full_name:
        raise HTTPException(status_code=400, detail="Lecturer full name is required.")

    temp_password = generate_temp_password()
    auth_ok, auth_message = upsert_supabase_lecturer(email, full_name, title, temp_password)

    existing = db.query(models.User).filter(models.User.email == email).first()
    if existing:
        if existing.role != models.UserRole.TEACHER:
            existing.role = models.UserRole.TEACHER
        existing.name = full_name
        existing.is_admin = is_admin
        lecturer = existing
    else:
        lecturer = models.User(
            id=uuid.uuid4(),
            email=email,
            name=full_name,
            role=models.UserRole.TEACHER,
            is_admin=is_admin,
        )
        db.add(lecturer)

    profile = db.query(models.LecturerProfile).filter(models.LecturerProfile.email == email).first()
    now = datetime.now(timezone.utc)
    if not profile:
        profile = models.LecturerProfile(
            id=uuid.uuid4(),
            email=email,
            title=title,
            full_name=full_name,
            must_change_password=True,
            created_at=now,
            updated_at=now,
        )
        db.add(profile)
    else:
        profile.title = title
        profile.full_name = full_name
        profile.must_change_password = True
        profile.updated_at = now

    ensure_local_credential(
        db,
        username=email,
        password=temp_password,
        role="TEACHER",
        user_email=email,
        force_update=True,
    )

    if is_admin:
        ensure_local_credential(
            db,
            username=email,
            password=temp_password,
            role="SUPERUSER",
            user_email=email,
            force_update=True,
        )

    db.commit()
    db.refresh(lecturer)

    return {
        "id": str(lecturer.id),
        "email": lecturer.email,
        "name": lecturer.name,
        "title": title,
        "fullName": full_name,
        "tempPassword": temp_password,
        "authProvisioned": auth_ok,
        "authMessage": auth_message,
        "role": lecturer.role.value,
        "isAdmin": bool(lecturer.is_admin),
        "message": "Lecturer added. Send temporary password to lecturer email and require password change at first login.",
    }

@app.get("/admin/auth/check")
def admin_auth_check(admin_context: dict = Depends(verify_admin_access)):
    return {
        "success": True,
        "mode": admin_context.get("mode"),
        "email": admin_context.get("email"),
        "isBootstrap": bool(admin_context.get("isBootstrap")),
        "isAdmin": bool(admin_context.get("isAdmin")),
    }

@app.post("/admin/lecturers/{lecturerId}/admin-access")
def admin_set_lecturer_admin_access(
    lecturerId: str,
    req: schemas.AdminUserAccessUpdate,
    _: dict = Depends(verify_admin_access),
    db: Session = Depends(get_db),
):
    lecturer_uuid = parse_uuid_or_400(lecturerId, "lecturerId")
    lecturer = db.query(models.User).filter(models.User.id == lecturer_uuid).first()
    if not lecturer:
        raise HTTPException(status_code=404, detail="Lecturer not found")
    if lecturer.role != models.UserRole.TEACHER:
        raise HTTPException(status_code=400, detail="Only lecturers can receive admin access")

    lecturer.is_admin = bool(req.is_admin)

    teacher_cred = (
        db.query(models.LocalCredential)
        .filter(
            models.LocalCredential.username == lecturer.email.lower(),
            models.LocalCredential.role == "TEACHER",
        )
        .first()
    )
    if teacher_cred:
        ensure_local_credential(
            db,
            username=lecturer.email.lower(),
            password="placeholder",
            role="SUPERUSER",
            user_email=lecturer.email.lower(),
        )
        super_cred = (
            db.query(models.LocalCredential)
            .filter(
                models.LocalCredential.username == lecturer.email.lower(),
                models.LocalCredential.role == "SUPERUSER",
            )
            .first()
        )
        if super_cred:
            super_cred.password_hash = teacher_cred.password_hash
            super_cred.active = bool(req.is_admin)
            super_cred.updated_at = datetime.now(timezone.utc)

    db.commit()
    return {
        "success": True,
        "message": "Admin access updated.",
        "id": str(lecturer.id),
        "isAdmin": bool(lecturer.is_admin),
    }

@app.delete("/admin/lecturers/{lecturerId}")
def admin_delete_lecturer(
    lecturerId: str,
    _: dict = Depends(verify_admin_access),
    db: Session = Depends(get_db),
):
    lecturer_uuid = parse_uuid_or_400(lecturerId, "lecturerId")
    lecturer = db.query(models.User).filter(models.User.id == lecturer_uuid).first()
    if not lecturer:
        raise HTTPException(status_code=404, detail="Lecturer not found")
    if lecturer.role != models.UserRole.TEACHER:
        raise HTTPException(status_code=400, detail="Selected user is not a lecturer")

    lecturer_email = lecturer.email.strip().lower()

    db.query(models.Course).filter(models.Course.lecturer_id == lecturer.id).update(
        {models.Course.lecturer_id: None},
        synchronize_session=False,
    )
    db.query(models.Session).filter(models.Session.lecturer_id == lecturer.id).update(
        {models.Session.lecturer_id: None},
        synchronize_session=False,
    )
    db.query(models.TeacherCourseAssignment).filter(
        models.TeacherCourseAssignment.teacher_email == lecturer_email
    ).delete(synchronize_session=False)
    db.query(models.PasswordResetRequest).filter(
        models.PasswordResetRequest.lecturer_email == lecturer_email
    ).delete(synchronize_session=False)
    db.query(models.LocalCredential).filter(
        models.LocalCredential.user_email == lecturer_email
    ).delete(synchronize_session=False)

    profile = db.query(models.LecturerProfile).filter(models.LecturerProfile.email == lecturer_email).first()
    if profile:
        db.delete(profile)

    db.delete(lecturer)
    db.commit()

    return {"success": True, "message": "Lecturer removed."}

@app.get("/admin/password-requests")
def admin_password_requests(_: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    items = (
        db.query(models.PasswordResetRequest)
        .order_by(models.PasswordResetRequest.created_at.desc())
        .all()
    )

    return [
        {
            "id": str(item.id),
            "lecturerEmail": item.lecturer_email,
            "lecturerTitle": item.lecturer_title,
            "lecturerName": item.lecturer_name,
            "status": item.status,
            "createdAt": item.created_at,
            "reviewedAt": item.reviewed_at,
            "adminNote": item.admin_note,
        }
        for item in items
    ]

@app.post("/admin/password-requests/{requestId}/issue-temp-password")
def admin_issue_temp_password(requestId: str, _: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    request_uuid = parse_uuid_or_400(requestId, "requestId")
    item = db.query(models.PasswordResetRequest).filter(models.PasswordResetRequest.id == request_uuid).first()
    if not item:
        raise HTTPException(status_code=404, detail="Password request not found")

    if item.status != "PENDING":
        raise HTTPException(status_code=400, detail="Password request already handled")

    temp_password = generate_temp_password()
    full_name = item.lecturer_name or item.lecturer_email.split("@")[0]
    title = item.lecturer_title or "Lecturer"
    auth_ok, auth_message = upsert_supabase_lecturer(item.lecturer_email, full_name, title, temp_password)

    profile = db.query(models.LecturerProfile).filter(models.LecturerProfile.email == item.lecturer_email).first()
    if profile:
        profile.must_change_password = True
        profile.updated_at = datetime.now(timezone.utc)

    ensure_local_credential(
        db,
        username=item.lecturer_email,
        password=temp_password,
        role="TEACHER",
        user_email=item.lecturer_email,
        force_update=True,
    )

    user = db.query(models.User).filter(models.User.email == item.lecturer_email).first()
    if user and user.is_admin:
        ensure_local_credential(
            db,
            username=item.lecturer_email,
            password=temp_password,
            role="SUPERUSER",
            user_email=item.lecturer_email,
            force_update=True,
        )

    item.status = "ISSUED"
    item.reviewed_at = datetime.now(timezone.utc)
    item.admin_note = "Temporary password issued by admin"
    db.commit()

    return {
        "success": True,
        "tempPassword": temp_password,
        "authProvisioned": auth_ok,
        "authMessage": auth_message,
        "message": "Temporary password generated. Send it to lecturer and require password change after login.",
    }

@app.get("/lecturers/password-policy")
def lecturer_password_policy(email: str, db: Session = Depends(get_db)):
    profile = db.query(models.LecturerProfile).filter(models.LecturerProfile.email == email.strip().lower()).first()
    return {
        "mustChangePassword": bool(profile.must_change_password) if profile else False,
        "title": profile.title if profile else None,
        "fullName": profile.full_name if profile else None,
    }

@app.post("/lecturers/password-policy/clear")
def lecturer_password_policy_clear(req: schemas.LecturerPasswordPolicyClear, db: Session = Depends(get_db)):
    email = req.email.strip().lower()
    profile = db.query(models.LecturerProfile).filter(models.LecturerProfile.email == email).first()
    if not profile:
        return {"success": True, "message": "No lecturer policy entry found."}

    profile.must_change_password = False
    profile.updated_at = datetime.now(timezone.utc)
    db.commit()
    return {"success": True, "message": "Password policy updated."}

@app.post("/lecturers/password-requests")
def lecturer_password_request(req: schemas.LecturerPasswordRequestCreate, db: Session = Depends(get_db)):
    email = req.email.strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Lecturer email is required.")

    profile = db.query(models.LecturerProfile).filter(models.LecturerProfile.email == email).first()
    user = db.query(models.User).filter(models.User.email == email).first()

    if not user or user.role != models.UserRole.TEACHER:
        raise HTTPException(status_code=404, detail="Lecturer account not found.")

    pending = (
        db.query(models.PasswordResetRequest)
        .filter(
            models.PasswordResetRequest.lecturer_email == email,
            models.PasswordResetRequest.status == "PENDING",
        )
        .first()
    )
    if pending:
        return {"success": True, "message": "Password request already pending admin review."}

    item = models.PasswordResetRequest(
        id=uuid.uuid4(),
        lecturer_email=email,
        lecturer_title=(profile.title if profile else "Lecturer"),
        lecturer_name=(profile.full_name if profile else user.name),
        status="PENDING",
        created_at=datetime.now(timezone.utc),
    )
    db.add(item)
    db.commit()

    return {"success": True, "message": "Password request sent to admin for validation."}

@app.post("/admin/course-requests/{requestId}/approve")
def admin_approve_course_request(requestId: str, _: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    request_uuid = parse_uuid_or_400(requestId, "requestId")
    item = db.query(models.CourseRequest).filter(models.CourseRequest.id == request_uuid).first()
    if not item:
        raise HTTPException(status_code=404, detail="Course request not found")

    if item.status != "PENDING":
        return {"success": True, "message": "Request already reviewed."}

    used_codes = {course.code for course in db.query(models.Course).all()}
    code = normalized_code(item.name, item.course_type, item.code or "", used_codes)

    existing_course = (
        db.query(models.Course)
        .filter(
            models.Course.name == item.name,
            models.Course.course_type == item.course_type,
            models.Course.semester == item.semester,
        )
        .first()
    )

    if not existing_course:
        db.add(
            models.Course(
                id=uuid.uuid4(),
                name=item.name,
                course_type=item.course_type,
                code=code,
                semester=item.semester,
                lecturer_id=None,
            )
        )

    item.status = "APPROVED"
    item.reviewed_at = datetime.now(timezone.utc)
    db.commit()

    return {"success": True, "message": "Course request approved and added to catalog."}

@app.delete("/admin/course-requests/{requestId}")
def admin_reject_course_request(requestId: str, _: dict = Depends(verify_admin_access), db: Session = Depends(get_db)):
    request_uuid = parse_uuid_or_400(requestId, "requestId")
    item = db.query(models.CourseRequest).filter(models.CourseRequest.id == request_uuid).first()
    if not item:
        raise HTTPException(status_code=404, detail="Course request not found")

    db.delete(item)
    db.commit()
    return {"success": True, "message": "Course request rejected and removed."}